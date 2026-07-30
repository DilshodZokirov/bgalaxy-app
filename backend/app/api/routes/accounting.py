import csv
import io
from datetime import date, datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func as sa_func
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.db.database import get_db
from app.models.accounting import Invoice, PayrollEntry, Transaction
from app.models.company import Company
from app.models.user import User
from app.schemas.accounting import (
    AccountingSummary,
    InvoiceCreate,
    InvoiceOut,
    InvoiceUpdate,
    PayrollCreate,
    PayrollOut,
    PeriodBucket,
    PeriodStats,
    PeriodTotals,
    TransactionCreate,
    TransactionOut,
    TransactionUpdate,
)
from app.services.invoice_docs import (
    build_invoice_pdf,
    calc_invoice_totals,
    ensure_invoice_income_tx,
    ensure_payroll_expense_tx,
    next_invoice_number,
    remove_invoice_income_tx,
    remove_payroll_expense_tx,
)
from app.services.permissions import require_permission

router = APIRouter(prefix="/companies/{company_id}/accounting", tags=["accounting"])


async def _check(db: AsyncSession, company_id: str, user_id) -> None:
    await require_permission(db, company_id, user_id, "manage_accounting")


def _month_bounds(month: str) -> tuple[date, date]:
    year, mon = (int(part) for part in month.split("-"))
    start = date(year, mon, 1)
    end = date(year + 1, 1, 1) if mon == 12 else date(year, mon + 1, 1)
    return start, end


def _refresh_overdue(inv: Invoice, today: date | None = None) -> bool:
    today = today or date.today()
    if inv.status in ("paid", "cancelled", "overdue"):
        return False
    if inv.due_date and inv.due_date < today and inv.status in ("draft", "sent"):
        inv.status = "overdue"
        return True
    return False


def _tx_out(tx: Transaction, creator_name: str | None) -> TransactionOut:
    return TransactionOut(
        id=tx.id,
        type=tx.type,
        category=tx.category,
        amount=float(tx.amount),
        description=tx.description,
        occurred_on=tx.occurred_on,
        created_at=tx.created_at,
        created_by_name=creator_name,
        source_invoice_id=tx.source_invoice_id,
        source_payroll_id=tx.source_payroll_id,
    )


def _invoice_out(inv: Invoice, creator_name: str | None) -> InvoiceOut:
    return InvoiceOut(
        id=inv.id,
        invoice_number=inv.invoice_number or f"INV-{str(inv.id)[:8].upper()}",
        client_name=inv.client_name,
        items=inv.items or [],
        vat_rate=float(inv.vat_rate or 0),
        subtotal_amount=float(inv.subtotal_amount or 0),
        vat_amount=float(inv.vat_amount or 0),
        total_amount=float(inv.total_amount),
        status=inv.status,
        issue_date=inv.issue_date,
        due_date=inv.due_date,
        created_at=inv.created_at,
        created_by_name=creator_name,
    )


def _payroll_out(entry: PayrollEntry, employee_name: str | None, creator_name: str | None) -> PayrollOut:
    return PayrollOut(
        id=entry.id,
        employee_id=entry.employee_id,
        employee_name=employee_name,
        period=entry.period,
        amount=float(entry.amount),
        status=entry.status,
        paid_at=entry.paid_at,
        created_by_name=creator_name,
    )


# ---------- Transactions (paginated, filterable, sortable, searchable) ----------


@router.get("/transactions")
async def list_transactions(
    company_id: str,
    page: int = 1,
    page_size: int = 5,
    search: str | None = None,
    type: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    sort_by: str = "occurred_on",
    sort_dir: str = "desc",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    query = (
        select(Transaction, User)
        .join(User, User.id == Transaction.created_by)
        .where(Transaction.company_id == company_id)
    )
    if type:
        query = query.where(Transaction.type == type)
    if date_from:
        query = query.where(Transaction.occurred_on >= date_from)
    if date_to:
        query = query.where(Transaction.occurred_on <= date_to)
    if search:
        like = f"%{search}%"
        query = query.where(or_(Transaction.category.ilike(like), Transaction.description.ilike(like)))

    sort_map = {"occurred_on": Transaction.occurred_on, "amount": Transaction.amount, "category": Transaction.category}
    sort_col = sort_map.get(sort_by, Transaction.occurred_on)
    query = query.order_by(sort_col.desc() if sort_dir == "desc" else sort_col.asc())

    count_result = await db.execute(select(sa_func.count()).select_from(query.subquery()))
    total = count_result.scalar_one()

    page = max(1, page)
    page_size = max(1, min(page_size, 100))
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    items = [_tx_out(tx, user.full_name) for tx, user in result.all()]

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.post("/transactions", response_model=TransactionOut, status_code=201)
async def create_transaction(
    company_id: str,
    payload: TransactionCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    if payload.type not in ("income", "expense"):
        raise HTTPException(status_code=400, detail="type 'income' yoki 'expense' bo'lishi kerak")
    tx = Transaction(
        company_id=company_id,
        type=payload.type,
        category=payload.category,
        amount=payload.amount,
        description=payload.description,
        occurred_on=payload.occurred_on,
        created_by=current_user.id,
    )
    db.add(tx)
    await db.commit()
    await db.refresh(tx)
    return _tx_out(tx, current_user.full_name)


@router.patch("/transactions/{transaction_id}", response_model=TransactionOut)
async def update_transaction(
    company_id: str,
    transaction_id: str,
    payload: TransactionUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    result = await db.execute(
        select(Transaction, User)
        .join(User, User.id == Transaction.created_by)
        .where(Transaction.id == transaction_id, Transaction.company_id == company_id)
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="Topilmadi")
    tx, creator = row
    data = payload.model_dump(exclude_unset=True)
    if "type" in data and data["type"] is not None and data["type"] not in ("income", "expense"):
        raise HTTPException(status_code=400, detail="type 'income' yoki 'expense' bo'lishi kerak")
    for key, value in data.items():
        setattr(tx, key, value)
    await db.commit()
    await db.refresh(tx)
    return _tx_out(tx, creator.full_name)


@router.delete("/transactions/{transaction_id}", status_code=204)
async def delete_transaction(
    company_id: str,
    transaction_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    result = await db.execute(
        select(Transaction).where(Transaction.id == transaction_id, Transaction.company_id == company_id)
    )
    tx = result.scalar_one_or_none()
    if tx is None:
        raise HTTPException(status_code=404, detail="Topilmadi")
    await db.delete(tx)
    await db.commit()


# ---------- Invoices ----------


@router.get("/invoices")
async def list_invoices(
    company_id: str,
    page: int = 1,
    page_size: int = 5,
    search: str | None = None,
    status: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    sort_by: str = "issue_date",
    sort_dir: str = "desc",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    query = select(Invoice, User).join(User, User.id == Invoice.created_by).where(Invoice.company_id == company_id)
    if status:
        query = query.where(Invoice.status == status)
    if date_from:
        query = query.where(Invoice.issue_date >= date_from)
    if date_to:
        query = query.where(Invoice.issue_date <= date_to)
    if search:
        like = f"%{search}%"
        query = query.where(or_(Invoice.client_name.ilike(like), Invoice.invoice_number.ilike(like)))

    sort_map = {
        "issue_date": Invoice.issue_date,
        "total_amount": Invoice.total_amount,
        "client_name": Invoice.client_name,
        "invoice_number": Invoice.invoice_number,
    }
    sort_col = sort_map.get(sort_by, Invoice.issue_date)
    query = query.order_by(sort_col.desc() if sort_dir == "desc" else sort_col.asc())

    count_result = await db.execute(select(sa_func.count()).select_from(query.subquery()))
    total = count_result.scalar_one()

    page = max(1, page)
    page_size = max(1, min(page_size, 100))
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    rows = result.all()
    changed = False
    today = date.today()
    for inv, _ in rows:
        if _refresh_overdue(inv, today):
            changed = True
    if changed:
        await db.commit()
    items = [_invoice_out(inv, user.full_name) for inv, user in rows]

    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.post("/invoices", response_model=InvoiceOut, status_code=201)
async def create_invoice(
    company_id: str,
    payload: InvoiceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    items = [item.model_dump() for item in payload.items]
    if not items:
        raise HTTPException(status_code=400, detail="Kamida bitta buyum kerak")
    subtotal, vat_amount, total = calc_invoice_totals(items, payload.vat_rate)
    number = await next_invoice_number(db, company_id, payload.issue_date)
    invoice = Invoice(
        company_id=company_id,
        invoice_number=number,
        client_name=payload.client_name.strip(),
        items=items,
        vat_rate=payload.vat_rate,
        subtotal_amount=subtotal,
        vat_amount=vat_amount,
        total_amount=total,
        issue_date=payload.issue_date,
        due_date=payload.due_date,
        created_by=current_user.id,
    )
    _refresh_overdue(invoice)
    db.add(invoice)
    await db.flush()
    if invoice.status == "paid":
        await ensure_invoice_income_tx(
            db, company_id=company_id, invoice=invoice, user_id=current_user.id
        )
    await db.commit()
    await db.refresh(invoice)
    return _invoice_out(invoice, current_user.full_name)


@router.get("/invoices/{invoice_id}/pdf")
async def download_invoice_pdf(
    company_id: str,
    invoice_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    result = await db.execute(
        select(Invoice, User)
        .join(User, User.id == Invoice.created_by)
        .where(Invoice.id == invoice_id, Invoice.company_id == company_id)
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="Topilmadi")
    invoice, creator = row
    company = (await db.execute(select(Company).where(Company.id == company_id))).scalar_one_or_none()
    if company is None:
        raise HTTPException(status_code=404, detail="Kompaniya topilmadi")
    pdf = build_invoice_pdf(company=company, invoice=invoice, creator_name=creator.full_name)
    filename = f"{(invoice.invoice_number or 'invoice').replace('/', '-')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.patch("/invoices/{invoice_id}", response_model=InvoiceOut)
async def update_invoice(
    company_id: str,
    invoice_id: str,
    payload: InvoiceUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    result = await db.execute(
        select(Invoice, User).join(User, User.id == Invoice.created_by).where(
            Invoice.id == invoice_id, Invoice.company_id == company_id
        )
    )
    row = result.first()
    if row is None:
        raise HTTPException(status_code=404, detail="Topilmadi")
    invoice, creator = row
    prev_status = invoice.status
    data = payload.model_dump(exclude_unset=True)

    if "status" in data and data["status"] is not None:
        if data["status"] not in ("draft", "sent", "paid", "overdue"):
            raise HTTPException(status_code=400, detail="Noto'g'ri status")
        invoice.status = data["status"]
    if "client_name" in data and data["client_name"] is not None:
        invoice.client_name = data["client_name"].strip()
    if "issue_date" in data and data["issue_date"] is not None:
        invoice.issue_date = data["issue_date"]
    if "due_date" in data:
        invoice.due_date = data["due_date"]
    if "vat_rate" in data and data["vat_rate"] is not None:
        invoice.vat_rate = data["vat_rate"]
    if "items" in data and data["items"] is not None:
        items = [item if isinstance(item, dict) else item.model_dump() for item in payload.items or []]
        if not items:
            raise HTTPException(status_code=400, detail="Kamida bitta buyum kerak")
        invoice.items = items

    items = invoice.items or []
    subtotal, vat_amount, total = calc_invoice_totals(items, float(invoice.vat_rate or 0))
    invoice.subtotal_amount = subtotal
    invoice.vat_amount = vat_amount
    invoice.total_amount = total
    _refresh_overdue(invoice)

    if invoice.status == "paid":
        await ensure_invoice_income_tx(
            db, company_id=company_id, invoice=invoice, user_id=current_user.id
        )
    elif prev_status == "paid" and invoice.status != "paid":
        await remove_invoice_income_tx(db, company_id=company_id, invoice_id=invoice.id)

    await db.commit()
    await db.refresh(invoice)
    return _invoice_out(invoice, creator.full_name)


@router.delete("/invoices/{invoice_id}", status_code=204)
async def delete_invoice(
    company_id: str,
    invoice_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    result = await db.execute(
        select(Invoice).where(Invoice.id == invoice_id, Invoice.company_id == company_id)
    )
    invoice = result.scalar_one_or_none()
    if invoice is None:
        raise HTTPException(status_code=404, detail="Topilmadi")
    await remove_invoice_income_tx(db, company_id=company_id, invoice_id=invoice.id)
    await db.delete(invoice)
    await db.commit()


# ---------- Payroll ----------


@router.get("/payroll")
async def list_payroll(
    company_id: str,
    page: int = 1,
    page_size: int = 5,
    search: str | None = None,
    period_from: str | None = None,
    period_to: str | None = None,
    sort_by: str = "period",
    sort_dir: str = "desc",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    Employee = User
    query = (
        select(PayrollEntry, Employee)
        .join(Employee, Employee.id == PayrollEntry.employee_id)
        .where(PayrollEntry.company_id == company_id)
    )
    if period_from:
        query = query.where(PayrollEntry.period >= period_from)
    if period_to:
        query = query.where(PayrollEntry.period <= period_to)
    if search:
        query = query.where(Employee.full_name.ilike(f"%{search}%"))

    sort_map = {"period": PayrollEntry.period, "amount": PayrollEntry.amount}
    sort_col = sort_map.get(sort_by, PayrollEntry.period)
    query = query.order_by(sort_col.desc() if sort_dir == "desc" else sort_col.asc())

    count_result = await db.execute(select(sa_func.count()).select_from(query.subquery()))
    total = count_result.scalar_one()

    page = max(1, page)
    page_size = max(1, min(page_size, 100))
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)

    # created_by names — batch lookup
    entries = result.all()
    creator_ids = {str(entry.created_by) for entry, _ in entries}
    creators = {}
    if creator_ids:
        creators_result = await db.execute(select(User).where(User.id.in_(creator_ids)))
        creators = {str(u.id): u.full_name for u in creators_result.scalars().all()}

    items = [_payroll_out(entry, employee.full_name, creators.get(str(entry.created_by))) for entry, employee in entries]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.post("/payroll", response_model=PayrollOut, status_code=201)
async def create_payroll(
    company_id: str,
    payload: PayrollCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    employee_result = await db.execute(select(User).where(User.id == str(payload.employee_id)))
    employee = employee_result.scalar_one_or_none()
    if employee is None:
        raise HTTPException(status_code=404, detail="Xodim topilmadi")

    entry = PayrollEntry(
        company_id=company_id,
        employee_id=payload.employee_id,
        period=payload.period,
        amount=payload.amount,
        created_by=current_user.id,
    )
    db.add(entry)
    await db.commit()
    await db.refresh(entry)
    return _payroll_out(entry, employee.full_name, current_user.full_name)


@router.patch("/payroll/{entry_id}/pay", response_model=PayrollOut)
async def mark_payroll_paid(
    company_id: str,
    entry_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    result = await db.execute(
        select(PayrollEntry).where(PayrollEntry.id == entry_id, PayrollEntry.company_id == company_id)
    )
    entry = result.scalar_one_or_none()
    if entry is None:
        raise HTTPException(status_code=404, detail="Topilmadi")

    entry.status = "paid"
    entry.paid_at = datetime.utcnow()
    employee = (
        await db.execute(select(User).where(User.id == str(entry.employee_id)))
    ).scalar_one_or_none()
    await ensure_payroll_expense_tx(
        db,
        company_id=company_id,
        entry=entry,
        employee_name=employee.full_name if employee else "Xodim",
        user_id=current_user.id,
    )
    await db.commit()

    user_ids = {str(entry.employee_id), str(entry.created_by)}
    users_result = await db.execute(select(User).where(User.id.in_(user_ids)))
    users_by_id = {str(u.id): u for u in users_result.scalars().all()}
    return _payroll_out(
        entry,
        users_by_id.get(str(entry.employee_id)) and users_by_id[str(entry.employee_id)].full_name,
        users_by_id.get(str(entry.created_by)) and users_by_id[str(entry.created_by)].full_name,
    )


@router.delete("/payroll/{entry_id}", status_code=204)
async def delete_payroll(
    company_id: str,
    entry_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    result = await db.execute(
        select(PayrollEntry).where(PayrollEntry.id == entry_id, PayrollEntry.company_id == company_id)
    )
    entry = result.scalar_one_or_none()
    if entry is None:
        raise HTTPException(status_code=404, detail="Topilmadi")
    await remove_payroll_expense_tx(db, company_id=company_id, payroll_id=entry.id)
    await db.delete(entry)
    await db.commit()


# ---------- Umumiy — month-scoped summary (stat cards) ----------


@router.get("/summary", response_model=AccountingSummary)
async def get_summary(
    company_id: str,
    month: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    month_start, month_end = _month_bounds(month)

    # Balans faqat kassa tranzaksiyalaridan (to'langan faktura/oylik linked tx orqali kiradi)
    tx_result = await db.execute(
        select(Transaction).where(
            Transaction.company_id == company_id,
            Transaction.occurred_on >= month_start,
            Transaction.occurred_on < month_end,
        )
    )
    transactions = tx_result.scalars().all()
    total_income = sum(float(t.amount) for t in transactions if t.type == "income")
    total_expense = sum(float(t.amount) for t in transactions if t.type == "expense")

    payroll_result = await db.execute(
        select(PayrollEntry).where(PayrollEntry.company_id == company_id, PayrollEntry.period == month)
    )
    total_payroll = sum(float(p.amount) for p in payroll_result.scalars().all())

    # Debitor qarz: yuborilgan / muddati o'tgan (kassaga tushmagan)
    inv_result = await db.execute(
        select(Invoice).where(
            Invoice.company_id == company_id,
            Invoice.status.in_(("draft", "sent", "overdue")),
        )
    )
    invoices = inv_result.scalars().all()
    changed = False
    today = date.today()
    for inv in invoices:
        if _refresh_overdue(inv, today):
            changed = True
    if changed:
        await db.commit()
    total_receivable = sum(
        float(inv.total_amount) for inv in invoices if inv.status in ("sent", "overdue", "draft")
    )

    return AccountingSummary(
        month=month,
        total_income=total_income,
        total_expense=total_expense,
        total_payroll=total_payroll,
        balance=total_income - total_expense,
        total_receivable=total_receivable,
    )


# ---------- Statistika — period-based chart data ----------


def _period_bounds(period: str) -> tuple[date, date, str]:
    today = datetime.utcnow().date()
    if period == "year":
        y, m = today.year - 1, today.month
        start = date(y, m, 1)
        bucket = "month"
    elif period == "6m":
        y, m = today.year, today.month - 5
        while m <= 0:
            m += 12
            y -= 1
        start = date(y, m, 1)
        bucket = "month"
    elif period == "3m":
        y, m = today.year, today.month - 2
        while m <= 0:
            m += 12
            y -= 1
        start = date(y, m, 1)
        bucket = "month"
    elif period == "1m":
        start = today.replace(day=1)
        bucket = "day"
    elif period == "1w":
        start = today - timedelta(days=6)
        bucket = "day"
    else:
        raise HTTPException(status_code=400, detail="Noto'g'ri davr")
    return start, today, bucket


@router.get("/stats", response_model=PeriodStats)
async def get_stats(
    company_id: str,
    period: str = "6m",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    start, end, bucket = _period_bounds(period)

    tx_result = await db.execute(
        select(Transaction).where(
            Transaction.company_id == company_id,
            Transaction.occurred_on >= start,
            Transaction.occurred_on <= end,
        )
    )
    all_tx = tx_result.scalars().all()

    payroll_result = await db.execute(select(PayrollEntry).where(PayrollEntry.company_id == company_id))
    all_payroll = payroll_result.scalars().all()

    def bucket_key(d: date) -> str:
        return d.strftime("%Y-%m") if bucket == "month" else d.strftime("%Y-%m-%d")

    keys: list[str] = []
    if bucket == "month":
        y, m = start.year, start.month
        while (y, m) <= (end.year, end.month):
            keys.append(f"{y}-{m:02d}")
            m += 1
            if m > 12:
                m = 1
                y += 1
    else:
        d = start
        while d <= end:
            keys.append(d.strftime("%Y-%m-%d"))
            d += timedelta(days=1)

    data = {k: {"income": 0.0, "expense": 0.0, "payroll": 0.0} for k in keys}

    for t in all_tx:
        k = bucket_key(t.occurred_on)
        if k in data:
            if t.type == "income":
                data[k]["income"] += float(t.amount)
            else:
                data[k]["expense"] += float(t.amount)

    # Oylik — faqat grafik uchun (balansga qayta ayirilmaydi; to'langan oylik expense tx da)
    if bucket == "month":
        for p in all_payroll:
            if p.period in data:
                data[p.period]["payroll"] += float(p.amount)

    buckets = [
        PeriodBucket(
            label=k,
            income=v["income"],
            expense=v["expense"],
            payroll=v["payroll"],
            balance=v["income"] - v["expense"],
        )
        for k, v in data.items()
    ]

    totals = PeriodTotals(
        total_income=sum(b.income for b in buckets),
        total_expense=sum(b.expense for b in buckets),
        total_payroll=sum(b.payroll for b in buckets),
        balance=sum(b.balance for b in buckets),
    )

    return PeriodStats(period=period, buckets=buckets, totals=totals)


# ---------- Report preview (JSON) + download (CSV), both date-range based ----------


@router.get("/yearly-summary")
async def get_yearly_summary(
    company_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Last 10 calendar years of income/expense/balance from cash transactions
    only (paid invoices/payroll appear via linked txs). Powers the Dashboard
    yearly comparison widget."""
    await _check(db, company_id, current_user.id)

    current_year = datetime.utcnow().year
    start_year = current_year - 9

    tx_result = await db.execute(
        select(Transaction).where(
            Transaction.company_id == company_id,
            Transaction.occurred_on >= date(start_year, 1, 1),
        )
    )
    all_tx = tx_result.scalars().all()

    years_data = {}
    for year in range(start_year, current_year + 1):
        income = sum(float(t.amount) for t in all_tx if t.type == "income" and t.occurred_on.year == year)
        expense = sum(float(t.amount) for t in all_tx if t.type == "expense" and t.occurred_on.year == year)
        years_data[year] = {
            "year": year,
            "income": income,
            "expense": expense,
            "balance": income - expense,
        }

    years = [years_data[y] for y in range(start_year, current_year + 1)]
    for i, entry in enumerate(years):
        prev = years[i - 1] if i > 0 else None
        entry["income_delta"] = entry["income"] - prev["income"] if prev else None
        entry["expense_delta"] = entry["expense"] - prev["expense"] if prev else None
        entry["balance_delta"] = entry["balance"] - prev["balance"] if prev else None

    return {"years": years}


@router.get("/report-data")
async def preview_report(
    company_id: str,
    date_from: date,
    date_to: date,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)

    tx_result = await db.execute(
        select(Transaction, User)
        .join(User, User.id == Transaction.created_by)
        .where(
            Transaction.company_id == company_id,
            Transaction.occurred_on >= date_from,
            Transaction.occurred_on <= date_to,
        )
        .order_by(Transaction.occurred_on)
    )
    invoice_result = await db.execute(
        select(Invoice, User)
        .join(User, User.id == Invoice.created_by)
        .where(Invoice.company_id == company_id, Invoice.issue_date >= date_from, Invoice.issue_date <= date_to)
    )
    payroll_result = await db.execute(
        select(PayrollEntry, User)
        .join(User, User.id == PayrollEntry.employee_id)
        .where(
            PayrollEntry.company_id == company_id,
            PayrollEntry.period >= date_from.strftime("%Y-%m"),
            PayrollEntry.period <= date_to.strftime("%Y-%m"),
        )
    )

    return {
        "date_from": str(date_from),
        "date_to": str(date_to),
        "transactions": [
            {
                "occurred_on": str(tx.occurred_on),
                "type": tx.type,
                "category": tx.category,
                "amount": float(tx.amount),
                "description": tx.description,
                "created_by_name": user.full_name,
            }
            for tx, user in tx_result.all()
        ],
        "invoices": [
            {
                "invoice_number": inv.invoice_number,
                "client_name": inv.client_name,
                "subtotal_amount": float(inv.subtotal_amount or 0),
                "vat_rate": float(inv.vat_rate or 0),
                "vat_amount": float(inv.vat_amount or 0),
                "total_amount": float(inv.total_amount),
                "issue_date": str(inv.issue_date),
                "status": inv.status,
                "created_by_name": user.full_name,
            }
            for inv, user in invoice_result.all()
        ],
        "payroll": [
            {
                "employee_name": employee.full_name,
                "period": entry.period,
                "amount": float(entry.amount),
                "status": entry.status,
            }
            for entry, employee in payroll_result.all()
        ],
    }


@router.get("/report")
async def download_report(
    company_id: str,
    date_from: date,
    date_to: date,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)

    tx_result = await db.execute(
        select(Transaction, User)
        .join(User, User.id == Transaction.created_by)
        .where(
            Transaction.company_id == company_id,
            Transaction.occurred_on >= date_from,
            Transaction.occurred_on <= date_to,
        )
        .order_by(Transaction.occurred_on)
    )
    invoice_result = await db.execute(
        select(Invoice, User)
        .join(User, User.id == Invoice.created_by)
        .where(Invoice.company_id == company_id, Invoice.issue_date >= date_from, Invoice.issue_date <= date_to)
    )
    payroll_result = await db.execute(
        select(PayrollEntry, User)
        .join(User, User.id == PayrollEntry.employee_id)
        .where(
            PayrollEntry.company_id == company_id,
            PayrollEntry.period >= date_from.strftime("%Y-%m"),
            PayrollEntry.period <= date_to.strftime("%Y-%m"),
        )
    )

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow([f"Hisobot — {date_from} / {date_to}"])
    writer.writerow([])
    writer.writerow(["TRANZAKSIYALAR"])
    writer.writerow(["Sana", "Turi", "Kategoriya", "Summasi", "Izoh", "Kim kiritdi"])
    for tx, user in tx_result.all():
        writer.writerow([tx.occurred_on, tx.type, tx.category, float(tx.amount), tx.description or "", user.full_name])

    writer.writerow([])
    writer.writerow(["HISOB-FAKTURALAR"])
    writer.writerow(["Raqam", "Mijoz", "Oraliq", "QQS %", "QQS", "Jami", "Sana", "Holati", "Kim yaratdi"])
    for inv, user in invoice_result.all():
        writer.writerow(
            [
                inv.invoice_number,
                inv.client_name,
                float(inv.subtotal_amount or 0),
                float(inv.vat_rate or 0),
                float(inv.vat_amount or 0),
                float(inv.total_amount),
                inv.issue_date,
                inv.status,
                user.full_name,
            ]
        )

    writer.writerow([])
    writer.writerow(["ISH HAQI"])
    writer.writerow(["Xodim", "Davr", "Summasi", "Holati"])
    for entry, employee in payroll_result.all():
        writer.writerow([employee.full_name, entry.period, float(entry.amount), entry.status])

    csv_content = buffer.getvalue()
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="hisobot-{date_from}_{date_to}.csv"'},
    )


# ---------- Excel report with selectable live formulas ----------

FORMULA_FUNCS = {
    "sum": ("SUM", "Jami"),
    "average": ("AVERAGE", "O'rtacha"),
    "max": ("MAX", "Eng ko'p"),
    "min": ("MIN", "Eng kam"),
    "count": ("COUNT", "Soni"),
    "median": ("MEDIAN", "Mediana"),
}

TX_TYPE_UZ = {"income": "Kirim", "expense": "Chiqim"}
INV_STATUS_UZ = {
    "draft": "Qoralama",
    "sent": "Yuborilgan",
    "paid": "To'langan",
    "overdue": "Muddati o'tgan",
}
PAY_STATUS_UZ = {"paid": "To'langan", "pending": "Kutilmoqda"}

_XL_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
_XL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_XL_TITLE_FONT = Font(bold=True, size=15, color="0F172A")
_XL_SUB_FONT = Font(size=11, color="475569")
_XL_SECTION_FILL = PatternFill("solid", fgColor="CCFBF1")
_XL_SECTION_FONT = Font(bold=True, size=11, color="115E59")
_XL_FORMULA_FILL = PatternFill("solid", fgColor="F0FDFA")
_XL_ZEBRA = PatternFill("solid", fgColor="F8FAFC")
_XL_MONEY = '#,##0.00'
_XL_THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _autosize(ws, min_width: int = 12, max_width: int = 36) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        length = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def _write_sheet(
    ws,
    title: str,
    period_label: str,
    headers: list[str],
    rows: list[list],
    amount_col_index: int,
    formula_keys: list[str],
) -> float:
    """Clean data table + optional Xulosa block with live Excel formulas."""
    col_count = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws["A1"] = title
    ws["A1"].font = _XL_TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws["A2"] = period_label
    ws["A2"].font = _XL_SUB_FONT

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _XL_HEADER_FONT
        cell.fill = _XL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _XL_THIN

    for offset, row in enumerate(rows):
        excel_row = header_row + 1 + offset
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = _XL_THIN
            cell.alignment = Alignment(vertical="center")
            if col_idx - 1 == amount_col_index and isinstance(value, (int, float)):
                cell.number_format = _XL_MONEY
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if offset % 2 == 1:
                cell.fill = _XL_ZEBRA

    data_start = header_row + 1
    data_end = header_row + len(rows)
    amount_total = 0.0
    for row in rows:
        value = row[amount_col_index]
        if isinstance(value, (int, float)):
            amount_total += float(value)

    if rows:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(col_count)}{data_end}"
        ws.freeze_panes = f"A{header_row + 1}"

    if rows and formula_keys:
        section_row = data_end + 2
        ws.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=col_count)
        section_cell = ws.cell(row=section_row, column=1, value="XULOSA (tanlangan formulalar)")
        section_cell.font = _XL_SECTION_FONT
        section_cell.fill = _XL_SECTION_FILL

        col_letter = get_column_letter(amount_col_index + 1)
        for i, key in enumerate(formula_keys):
            func, label = FORMULA_FUNCS.get(key, (None, None))
            if not func:
                continue
            excel_row = section_row + 1 + i
            label_cell = ws.cell(row=excel_row, column=1, value=label)
            label_cell.font = Font(bold=True, color="0F172A")
            label_cell.fill = _XL_FORMULA_FILL
            label_cell.border = _XL_THIN

            for col_idx in range(2, col_count + 1):
                filler = ws.cell(row=excel_row, column=col_idx, value="")
                filler.fill = _XL_FORMULA_FILL
                filler.border = _XL_THIN

            formula_cell = ws.cell(
                row=excel_row,
                column=amount_col_index + 1,
                value=f"={func}({col_letter}{data_start}:{col_letter}{data_end})",
            )
            formula_cell.font = Font(bold=True, color="0F766E")
            formula_cell.number_format = _XL_MONEY if key != "count" else "0"
            formula_cell.fill = _XL_FORMULA_FILL
            formula_cell.border = _XL_THIN
            formula_cell.alignment = Alignment(horizontal="right")

    _autosize(ws)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[header_row].height = 20
    return amount_total


@router.get("/report-excel")
async def download_report_excel(
    company_id: str,
    date_from: date,
    date_to: date,
    formulas: str = "sum",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _check(db, company_id, current_user.id)
    # Preserve user order, drop unknown/duplicates — avoids messy duplicate blocks.
    seen: set[str] = set()
    formula_keys: list[str] = []
    for raw in formulas.split(","):
        key = raw.strip()
        if key in FORMULA_FUNCS and key not in seen:
            seen.add(key)
            formula_keys.append(key)

    company = (
        await db.execute(select(Company).where(Company.id == company_id))
    ).scalar_one_or_none()
    company_name = company.name if company else "Kompaniya"
    period_label = f"Davr: {date_from} — {date_to}"

    tx_result = await db.execute(
        select(Transaction, User)
        .join(User, User.id == Transaction.created_by)
        .where(
            Transaction.company_id == company_id,
            Transaction.occurred_on >= date_from,
            Transaction.occurred_on <= date_to,
        )
        .order_by(Transaction.occurred_on, Transaction.created_at)
    )
    invoice_result = await db.execute(
        select(Invoice, User)
        .join(User, User.id == Invoice.created_by)
        .where(Invoice.company_id == company_id, Invoice.issue_date >= date_from, Invoice.issue_date <= date_to)
        .order_by(Invoice.issue_date, Invoice.created_at)
    )
    payroll_result = await db.execute(
        select(PayrollEntry, User)
        .join(User, User.id == PayrollEntry.employee_id)
        .where(
            PayrollEntry.company_id == company_id,
            PayrollEntry.period >= date_from.strftime("%Y-%m"),
            PayrollEntry.period <= date_to.strftime("%Y-%m"),
        )
        .order_by(PayrollEntry.period, User.full_name)
    )

    tx_rows = [
        [
            str(tx.occurred_on),
            TX_TYPE_UZ.get(tx.type, tx.type),
            tx.category,
            float(tx.amount),
            tx.description or "",
            user.full_name,
        ]
        for tx, user in tx_result.all()
    ]
    inv_rows = [
        [
            inv.invoice_number,
            inv.client_name,
            float(inv.subtotal_amount or 0),
            float(inv.vat_rate or 0),
            float(inv.vat_amount or 0),
            float(inv.total_amount),
            str(inv.issue_date),
            INV_STATUS_UZ.get(inv.status, inv.status),
            user.full_name,
        ]
        for inv, user in invoice_result.all()
    ]
    payroll_rows = [
        [
            employee.full_name,
            entry.period,
            float(entry.amount),
            PAY_STATUS_UZ.get(entry.status, entry.status),
        ]
        for entry, employee in payroll_result.all()
    ]

    wb = Workbook()

    # --- Cover / Umumiy ---
    ws_cover = wb.active
    ws_cover.title = "Umumiy"
    ws_cover["A1"] = "BG — Buxgalteriya hisoboti"
    ws_cover["A1"].font = _XL_TITLE_FONT
    ws_cover["A2"] = company_name
    ws_cover["A2"].font = Font(bold=True, size=13, color="0F766E")
    ws_cover["A3"] = period_label
    ws_cover["A3"].font = _XL_SUB_FONT
    ws_cover["A4"] = f"Yaratilgan: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    ws_cover["A4"].font = _XL_SUB_FONT

    ws_cover["A6"] = "Bo'lim"
    ws_cover["B6"] = "Yozuvlar"
    ws_cover["C6"] = "Summa"
    for cell in (ws_cover["A6"], ws_cover["B6"], ws_cover["C6"]):
        cell.font = _XL_HEADER_FONT
        cell.fill = _XL_HEADER_FILL
        cell.border = _XL_THIN
        cell.alignment = Alignment(horizontal="center")

    section_stats = [
        ("Tranzaksiyalar", len(tx_rows), sum(r[3] for r in tx_rows)),
        ("Hisob-fakturalar", len(inv_rows), sum(r[5] for r in inv_rows)),
        ("Ish haqi", len(payroll_rows), sum(r[2] for r in payroll_rows)),
    ]
    for i, (name, count, total) in enumerate(section_stats, start=7):
        ws_cover.cell(row=i, column=1, value=name).border = _XL_THIN
        count_cell = ws_cover.cell(row=i, column=2, value=count)
        count_cell.border = _XL_THIN
        count_cell.alignment = Alignment(horizontal="center")
        total_cell = ws_cover.cell(row=i, column=3, value=total)
        total_cell.border = _XL_THIN
        total_cell.number_format = _XL_MONEY

    ws_cover["A11"] = "Tanlangan formulalar"
    ws_cover["A11"].font = Font(bold=True, color="115E59")
    if formula_keys:
        for i, key in enumerate(formula_keys):
            _, label = FORMULA_FUNCS[key]
            cell = ws_cover.cell(row=12 + i, column=1, value=f"• {label} ({FORMULA_FUNCS[key][0]})")
            cell.font = Font(color="0F172A")
        note_row = 12 + len(formula_keys) + 1
    else:
        ws_cover["A12"] = "• Formulalar tanlanmagan (faqat ma'lumotlar)"
        note_row = 14
    ws_cover.cell(
        row=note_row,
        column=1,
        value="Har bir bo'lim varag'ida ma'lumotlar alohida, formulalar esa pastdagi «Xulosa» qismida tartibli joylashgan.",
    ).font = _XL_SUB_FONT
    _autosize(ws_cover, min_width=14, max_width=55)

    ws_tx = wb.create_sheet("Tranzaksiyalar")
    _write_sheet(
        ws_tx,
        "Tranzaksiyalar",
        period_label,
        ["Sana", "Turi", "Kategoriya", "Summasi", "Izoh", "Kim kiritdi"],
        tx_rows,
        3,
        formula_keys,
    )

    ws_inv = wb.create_sheet("Hisob-fakturalar")
    _write_sheet(
        ws_inv,
        "Hisob-fakturalar",
        period_label,
        ["Raqam", "Mijoz", "Oraliq", "QQS %", "QQS", "Jami", "Sana", "Holati", "Kim yaratdi"],
        inv_rows,
        5,
        formula_keys,
    )

    ws_payroll = wb.create_sheet("Ish haqi")
    _write_sheet(
        ws_payroll,
        "Ish haqi",
        period_label,
        ["Xodim", "Davr", "Summasi", "Holati"],
        payroll_rows,
        2,
        formula_keys,
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="hisobot-{date_from}_{date_to}.xlsx"'},
    )
