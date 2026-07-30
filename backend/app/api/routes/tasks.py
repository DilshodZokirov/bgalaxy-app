import io
import os
import uuid as uuid_lib
from datetime import date, datetime, timedelta

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    UploadFile,
    WebSocket,
    WebSocketDisconnect,
)
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func as sa_func
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.core.security import decode_access_token
from app.db.database import async_session, get_db
from app.models.company import TeamMembership
from app.models.notification import Notification
from app.models.role import Role
from app.models.task import Task, TaskComment, task_points
from app.models.user import User
from app.schemas.task import TaskCommentOut, TaskCreate, TaskOut, TaskUpdate
from app.services.connection_manager import task_board_manager
from app.services.notify import ping_notifications
from app.services.permissions import get_permissions, require_permission

router = APIRouter(prefix="/companies/{company_id}/tasks", tags=["tasks"])

ACTIVE_STATUSES = ("todo", "in_progress", "testing")
FINAL_STATUSES = ("accepted", "rejected", "failed")
PRIORITY_LABELS = {"hard": "Qiyin", "medium": "O'rtacha", "easy": "Oson"}
STATUS_LABELS = {
    "todo": "Bajarilmagan",
    "in_progress": "Ishda",
    "testing": "Tekshiruvda",
    "accepted": "Bajarilgan",
    "rejected": "Qabul qilinmagan",
    "failed": "Bajarilmagan",
}

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)
MAX_COMMENT_FILE_BYTES = 20 * 1024 * 1024


async def _names_map(db: AsyncSession, user_ids: set) -> dict:
    user_ids = {uid for uid in user_ids if uid}
    if not user_ids:
        return {}
    result = await db.execute(select(User).where(User.id.in_(user_ids)))
    return {str(u.id): u.full_name for u in result.scalars().all()}


async def _comment_stats(db: AsyncSession, task_ids: list) -> dict[str, tuple[int, int]]:
    if not task_ids:
        return {}
    result = await db.execute(
        select(
            TaskComment.task_id,
            sa_func.count(TaskComment.id),
            sa_func.count(TaskComment.file_url),
        )
        .where(TaskComment.task_id.in_(task_ids))
        .group_by(TaskComment.task_id)
    )
    return {str(tid): (int(cc or 0), int(fc or 0)) for tid, cc, fc in result.all()}


def _to_out(task: Task, names: dict, comment_count: int = 0, file_count: int = 0) -> TaskOut:
    return TaskOut(
        id=task.id,
        title=task.title,
        description=task.description,
        priority=task.priority,
        status=task.status,
        assignee_id=task.assignee_id,
        assignee_name=names.get(str(task.assignee_id)),
        created_by=task.created_by,
        created_by_name=names.get(str(task.created_by)),
        checked_by=task.checked_by,
        checked_by_name=names.get(str(task.checked_by)) if task.checked_by else ("Hech kim" if task.status == "failed" else None),
        start_date=task.start_date,
        due_date=task.due_date,
        completed_at=task.completed_at,
        points=task_points(task.status, task.priority),
        comment_count=comment_count,
        file_count=file_count,
        created_at=task.created_at,
        updated_at=task.updated_at,
    )


async def _tasks_out(db: AsyncSession, tasks: list[Task]) -> list[TaskOut]:
    if not tasks:
        return []
    names = await _names_map(
        db,
        {str(t.assignee_id) for t in tasks}
        | {str(t.created_by) for t in tasks}
        | {str(t.checked_by or "") for t in tasks},
    )
    stats = await _comment_stats(db, [t.id for t in tasks])
    out = []
    for t in tasks:
        cc, fc = stats.get(str(t.id), (0, 0))
        out.append(_to_out(t, names, cc, fc))
    return out


async def _broadcast_board(company_id: str, event_type: str, **extra) -> None:
    await task_board_manager.broadcast(
        str(company_id),
        {"type": event_type, "company_id": str(company_id), **extra},
    )


async def _require_member(db: AsyncSession, company_id: str, user_id) -> None:
    """Owner or team member may subscribe to the company task board."""
    from app.models.company import Company

    company_result = await db.execute(select(Company).where(Company.id == company_id))
    company = company_result.scalar_one_or_none()
    if company and str(company.owner_id) == str(user_id):
        return

    result = await db.execute(
        select(TeamMembership).where(
            TeamMembership.company_id == company_id,
            TeamMembership.user_id == user_id,
        )
    )
    if result.scalar_one_or_none() is None:
        raise HTTPException(status_code=403, detail="Siz bu kompaniya a'zosi emassiz")


async def _get_visible_task(
    db: AsyncSession, company_id: str, task_id: str, current_user: User
) -> tuple[Task, bool]:
    result = await db.execute(select(Task).where(Task.id == task_id, Task.company_id == company_id))
    task = result.scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Vazifa topilmadi")
    perms = await get_permissions(db, company_id, current_user.id)
    is_pm = perms["is_owner"] or perms["permissions"].get("manage_tasks", False)
    if not is_pm and str(task.assignee_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Bu vazifa sizga tegishli emas")
    return task, is_pm


def _comment_to_out(comment: TaskComment, author_name: str | None) -> TaskCommentOut:
    return TaskCommentOut(
        id=comment.id,
        task_id=comment.task_id,
        author_id=comment.author_id,
        author_name=author_name,
        content=comment.content,
        file_url=comment.file_url,
        file_name=comment.file_name,
        created_at=comment.created_at,
    )


async def _expire_overdue(db: AsyncSession, company_id: str) -> None:
    today = date.today()
    result = await db.execute(
        select(Task).where(
            Task.company_id == company_id,
            Task.status.in_(ACTIVE_STATUSES),
            Task.due_date < today,
        )
    )
    overdue = result.scalars().all()
    if not overdue:
        return
    for task in overdue:
        task.status = "failed"
        task.completed_at = task.due_date
        task.checked_by = None
        db.add(
            Notification(
                user_id=task.assignee_id,
                type="task_update",
                message=f"Afsuski, \"{task.title}\" topshirig'ingiz muddati o'tib ketgani uchun muvaffaqiyatsiz yakunlandi.",
                company_id=company_id,
            )
        )
    await db.commit()
    for task in overdue:
        await ping_notifications(task.assignee_id)
    await _broadcast_board(company_id, "tasks_changed", reason="expired")


async def _notify_pms(db: AsyncSession, company_id: str, message: str, exclude_user_id) -> set[str]:
    from app.models.company import Company

    company_result = await db.execute(select(Company).where(Company.id == company_id))
    company = company_result.scalar_one_or_none()

    result = await db.execute(
        select(User, TeamMembership)
        .join(TeamMembership, TeamMembership.user_id == User.id)
        .where(TeamMembership.company_id == company_id)
    )
    rows = result.all()
    roles_result = await db.execute(select(Role).where(Role.company_id == company_id))
    pm_role_ids = {r.id for r in roles_result.scalars().all() if r.permissions.get("manage_tasks")}

    notified: set[str] = set()
    for user, membership in rows:
        if str(user.id) == str(exclude_user_id):
            continue
        is_owner = company and str(user.id) == str(company.owner_id)
        if is_owner or membership.role_id in pm_role_ids:
            if str(user.id) not in notified:
                db.add(Notification(user_id=user.id, type="task_update", message=message, company_id=company_id))
                notified.add(str(user.id))
    return notified


async def _resolve_targets(db: AsyncSession, company_id: str, target_type: str, target_ids: list) -> list[str]:
    if target_type == "everyone":
        result = await db.execute(
            select(TeamMembership.user_id).where(TeamMembership.company_id == company_id)
        )
        return [str(uid) for uid in result.scalars().all()]
    if target_type == "role":
        if not target_ids:
            raise HTTPException(status_code=400, detail="Lavozim tanlanmagan")
        result = await db.execute(
            select(TeamMembership.user_id).where(
                TeamMembership.company_id == company_id, TeamMembership.role_id == str(target_ids[0])
            )
        )
        ids = [str(uid) for uid in result.scalars().all()]
        if not ids:
            raise HTTPException(status_code=400, detail="Bu lavozimda hech kim yo'q")
        return ids
    if target_type == "users":
        if not target_ids:
            raise HTTPException(status_code=400, detail="Kamida bitta odam tanlang")
        return [str(uid) for uid in target_ids]
    raise HTTPException(status_code=400, detail="Noto'g'ri target_type")


@router.get("", response_model=list[TaskOut])
async def list_tasks(
    company_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _expire_overdue(db, company_id)
    perms = await get_permissions(db, company_id, current_user.id)
    is_pm = perms["is_owner"] or perms["permissions"].get("manage_tasks", False)

    query = select(Task).where(Task.company_id == company_id, Task.status.in_(ACTIVE_STATUSES))
    if not is_pm:
        query = query.where(Task.assignee_id == current_user.id)
    result = await db.execute(query.order_by(Task.created_at))
    tasks = result.scalars().all()
    return await _tasks_out(db, tasks)


@router.post("", response_model=list[TaskOut], status_code=201)
async def create_task(
    company_id: str,
    payload: TaskCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await require_permission(db, company_id, current_user.id, "manage_tasks")
    if payload.priority not in ("hard", "medium", "easy"):
        raise HTTPException(status_code=400, detail="Noto'g'ri daraja")

    assignee_ids = await _resolve_targets(db, company_id, payload.target_type, payload.target_ids)
    assignee_ids = [uid for uid in assignee_ids if uid != str(current_user.id)]
    if not assignee_ids:
        raise HTTPException(status_code=400, detail="O'zingizga vazifa bera olmaysiz — boshqa odam yoki guruh tanlang")
    today = date.today()

    created_tasks: list[Task] = []
    for uid in assignee_ids:
        task = Task(
            company_id=company_id,
            title=payload.title,
            description=payload.description,
            priority=payload.priority,
            status="todo",
            assignee_id=uid,
            created_by=current_user.id,
            start_date=today,
            due_date=payload.due_date,
        )
        db.add(task)
        created_tasks.append(task)
        db.add(
            Notification(
                user_id=uid,
                type="task_assigned",
                message=f"{current_user.full_name} sizga yangi vazifa berdi: \"{payload.title}\"",
                company_id=company_id,
            )
        )

    await db.commit()
    for t in created_tasks:
        await db.refresh(t)
    for uid in assignee_ids:
        await ping_notifications(uid)

    out = await _tasks_out(db, created_tasks)
    await _broadcast_board(
        company_id,
        "tasks_changed",
        reason="created",
        task_ids=[str(t.id) for t in created_tasks],
    )
    return out


@router.patch("/{task_id}", response_model=TaskOut)
async def update_task(
    company_id: str,
    task_id: str,
    payload: TaskUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Task).where(Task.id == task_id, Task.company_id == company_id))
    task = result.scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Vazifa topilmadi")

    perms = await get_permissions(db, company_id, current_user.id)
    is_pm = perms["is_owner"] or perms["permissions"].get("manage_tasks", False)
    review_notified_id = None
    pm_notified: set[str] = set()

    if is_pm:
        if payload.title is not None:
            task.title = payload.title
        if payload.description is not None:
            task.description = payload.description
        if payload.priority is not None:
            if payload.priority not in ("hard", "medium", "easy"):
                raise HTTPException(status_code=400, detail="Noto'g'ri daraja")
            task.priority = payload.priority
        if payload.due_date is not None:
            task.due_date = payload.due_date

        if payload.status is not None and payload.status != task.status:
            if payload.status in ("accepted", "rejected"):
                task.status = payload.status
                task.checked_by = current_user.id
                task.completed_at = date.today()
                points = task_points(task.status, task.priority)
                db.add(
                    Notification(
                        user_id=task.assignee_id,
                        type="task_update",
                        message=(
                            (
                                f"\"{task.title}\" topshirig'ingiz muvaffaqiyatli tekshirildi! "
                                if task.status == "accepted"
                                else f"Afsuski, \"{task.title}\" bajargan topshirig'ingiz muvaffaqiyatsiz yakunlandi. "
                            )
                            + f"({points:+d} ball)."
                        ),
                        company_id=company_id,
                    )
                )
                review_notified_id = task.assignee_id
            elif payload.status in ACTIVE_STATUSES:
                task.status = payload.status
            else:
                raise HTTPException(status_code=400, detail="Noto'g'ri holat")
    else:
        if str(task.assignee_id) != str(current_user.id):
            raise HTTPException(status_code=403, detail="Bu vazifa sizga tegishli emas")
        if payload.status is None or payload.status not in ACTIVE_STATUSES:
            raise HTTPException(status_code=403, detail="Sizda faqat holatni (todo/ishda/tekshiruvda) o'zgartirish huquqi bor")
        if task.status != payload.status:
            task.status = payload.status
            if payload.status == "testing":
                pm_notified = await _notify_pms(
                    db,
                    company_id,
                    f"\"{task.title}\" vazifasi tekshiruvga yuborildi ({current_user.full_name})",
                    current_user.id,
                )

    await db.commit()
    await db.refresh(task)
    if review_notified_id:
        await ping_notifications(review_notified_id)
    for uid in pm_notified:
        await ping_notifications(uid)

    out = (await _tasks_out(db, [task]))[0]
    await _broadcast_board(
        company_id,
        "tasks_changed",
        reason="updated",
        task_id=str(task.id),
        status=task.status,
    )
    return out


@router.delete("/{task_id}", status_code=204)
async def delete_task(
    company_id: str,
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await require_permission(db, company_id, current_user.id, "manage_tasks")
    result = await db.execute(select(Task).where(Task.id == task_id, Task.company_id == company_id))
    task = result.scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=404, detail="Vazifa topilmadi")
    await db.delete(task)
    await db.commit()
    await _broadcast_board(company_id, "tasks_changed", reason="deleted", task_id=str(task_id))


@router.get("/monthly-champion")
async def monthly_champion(
    company_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Top scorer for the *previous* calendar month — Statistika → Jamoa
    "O'tgan oyning eng faol ishchisi".

    Goes live from August 2026 (when August opens, July is scored). Until
    then always returns null. Also null if nobody completed
    (accepted/rejected) tasks in that previous month."""
    result = await db.execute(
        select(TeamMembership.user_id).where(TeamMembership.company_id == company_id)
    )
    member_ids = {str(uid) for uid in result.scalars().all()}
    if str(current_user.id) not in member_ids:
        raise HTTPException(status_code=403, detail="Siz bu kompaniya a'zosi emassiz")

    today = date.today()
    # Launch: August 2026 → first visible champion month is July 2026.
    if today < date(2026, 8, 1):
        return None

    this_month_start = today.replace(day=1)
    prev_month_end = this_month_start - timedelta(days=1)
    prev_month_start = prev_month_end.replace(day=1)

    task_result = await db.execute(
        select(Task).where(
            Task.company_id == company_id,
            Task.status.in_(("accepted", "rejected")),
            Task.completed_at >= prev_month_start,
            Task.completed_at <= prev_month_end,
        )
    )
    tasks = task_result.scalars().all()
    if not tasks:
        return None

    scores: dict[str, dict] = {}
    for t in tasks:
        uid = str(t.assignee_id)
        entry = scores.setdefault(uid, {"score": 0, "accepted": 0, "rejected": 0})
        entry["score"] += task_points(t.status, t.priority)
        entry["accepted" if t.status == "accepted" else "rejected"] += 1

    top_uid = max(scores, key=lambda uid: scores[uid]["score"])
    top = scores[top_uid]

    user_result = await db.execute(select(User).where(User.id == top_uid))
    top_user = user_result.scalar_one_or_none()
    if top_user is None:
        return None

    role_result = await db.execute(
        select(Role)
        .join(TeamMembership, TeamMembership.role_id == Role.id)
        .where(TeamMembership.company_id == company_id, TeamMembership.user_id == top_uid)
    )
    role = role_result.scalar_one_or_none()

    return {
        "user_id": top_uid,
        "full_name": top_user.full_name,
        "avatar_url": top_user.avatar_url,
        "role_name": role.name if role else None,
        "score": top["score"],
        "accepted": top["accepted"],
        "rejected": top["rejected"],
        "month": prev_month_start.strftime("%Y-%m"),
        "month_start": prev_month_start.isoformat(),
        "month_end": prev_month_end.isoformat(),
    }


@router.get("/leaderboard")
async def task_leaderboard(
    company_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Visible to any company member — an aggregate score per person based on
    their accepted/rejected tasks, sorted highest first."""
    result = await db.execute(
        select(TeamMembership.user_id).where(TeamMembership.company_id == company_id)
    )
    member_ids = {str(uid) for uid in result.scalars().all()}
    if str(current_user.id) not in member_ids:
        raise HTTPException(status_code=403, detail="Siz bu kompaniya a'zosi emassiz")

    task_result = await db.execute(
        select(Task).where(Task.company_id == company_id, Task.status.in_(("accepted", "rejected")))
    )
    tasks = task_result.scalars().all()

    scores: dict[str, dict] = {}
    for t in tasks:
        uid = str(t.assignee_id)
        entry = scores.setdefault(uid, {"score": 0, "accepted": 0, "rejected": 0})
        entry["score"] += task_points(t.status, t.priority)
        entry["accepted" if t.status == "accepted" else "rejected"] += 1

    all_names = await _names_map(db, member_ids)
    rows = [
        {
            "user_id": uid,
            "full_name": all_names.get(uid, "Noma'lum"),
            "score": data["score"],
            "accepted": data["accepted"],
            "rejected": data["rejected"],
        }
        for uid, data in scores.items()
    ]
    # Include members with zero completed tasks too, at the bottom.
    for uid, name in all_names.items():
        if uid not in scores:
            rows.append({"user_id": uid, "full_name": name, "score": 0, "accepted": 0, "rejected": 0})

    rows.sort(key=lambda r: r["score"], reverse=True)
    return rows


@router.get("/history")
async def task_history(
    company_id: str,
    page: int = 1,
    page_size: int = 10,
    search: str | None = None,
    priority: str | None = None,
    status: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    sort_by: str = "completed_at",
    sort_dir: str = "desc",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _expire_overdue(db, company_id)
    perms = await get_permissions(db, company_id, current_user.id)
    is_pm = perms["is_owner"] or perms["permissions"].get("manage_tasks", False)

    query = select(Task).where(Task.company_id == company_id, Task.status.in_(FINAL_STATUSES))
    if not is_pm:
        query = query.where(Task.assignee_id == current_user.id)
    if search:
        query = query.where(Task.title.ilike(f"%{search}%"))
    if priority:
        query = query.where(Task.priority == priority)
    if status:
        query = query.where(Task.status == status)
    if date_from:
        query = query.where(Task.completed_at >= date_from)
    if date_to:
        query = query.where(Task.completed_at <= date_to)

    sort_map = {"completed_at": Task.completed_at, "due_date": Task.due_date, "priority": Task.priority, "title": Task.title}
    sort_col = sort_map.get(sort_by, Task.completed_at)
    query = query.order_by(sort_col.desc() if sort_dir == "desc" else sort_col.asc())

    count_result = await db.execute(select(sa_func.count()).select_from(query.subquery()))
    total = count_result.scalar_one()

    page = max(1, page)
    page_size = max(1, min(page_size, 100))
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    tasks = result.scalars().all()
    items = await _tasks_out(db, tasks)
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/history/report-excel")
async def download_history_excel(
    company_id: str,
    date_from: date,
    date_to: date,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    perms = await get_permissions(db, company_id, current_user.id)
    is_pm = perms["is_owner"] or perms["permissions"].get("manage_tasks", False)

    query = select(Task).where(
        Task.company_id == company_id,
        Task.status.in_(FINAL_STATUSES),
        Task.completed_at >= date_from,
        Task.completed_at <= date_to,
    )
    if not is_pm:
        query = query.where(Task.assignee_id == current_user.id)
    result = await db.execute(query.order_by(Task.completed_at))
    tasks = result.scalars().all()

    names = await _names_map(
        db, {str(t.assignee_id) for t in tasks} | {str(t.created_by) for t in tasks} | {str(t.checked_by or "") for t in tasks}
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Vazifalar tarixi"
    headers = ["Yaratilgan", "Yaratuvchi", "Mazmuni", "Daraja", "Muddat", "Kimga", "Tugatilgan", "Tekshirgan", "Holati", "Ball"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="1E2938", end_color="1E2938", fill_type="solid")

    for t in tasks:
        ws.append([
            str(t.created_at.date()),
            names.get(str(t.created_by), ""),
            t.title,
            PRIORITY_LABELS.get(t.priority, t.priority),
            str(t.due_date),
            names.get(str(t.assignee_id), ""),
            str(t.completed_at) if t.completed_at else "",
            names.get(str(t.checked_by), "Hech kim") if t.checked_by else "Hech kim",
            STATUS_LABELS.get(t.status, t.status),
            task_points(t.status, t.priority),
        ])

    for column_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="vazifalar-{date_from}_{date_to}.xlsx"'},
    )


async def _list_comments_impl(db: AsyncSession, company_id: str, task_id: str, current_user: User):
    await _get_visible_task(db, company_id, task_id, current_user)
    result = await db.execute(
        select(TaskComment, User)
        .join(User, User.id == TaskComment.author_id)
        .where(TaskComment.task_id == task_id, TaskComment.company_id == company_id)
        .order_by(TaskComment.created_at)
    )
    return [_comment_to_out(c, u.full_name) for c, u in result.all()]


async def _create_comment_impl(
    db: AsyncSession,
    company_id: str,
    task_id: str,
    current_user: User,
    content: str,
    file: UploadFile | None,
):
    await _get_visible_task(db, company_id, task_id, current_user)

    file_url = None
    file_name = None
    if file is not None and file.filename:
        data = await file.read()
        if len(data) > MAX_COMMENT_FILE_BYTES:
            raise HTTPException(status_code=400, detail="Fayl hajmi juda katta (20 MB dan kichik bo'lsin)")
        ext = os.path.splitext(file.filename)[1]
        stored_name = f"{uuid_lib.uuid4().hex}{ext}"
        with open(os.path.join(UPLOAD_DIR, stored_name), "wb") as f:
            f.write(data)
        file_url = f"/uploads/{stored_name}"
        file_name = file.filename

    text = (content or "").strip()
    if not text and not file_url:
        raise HTTPException(status_code=400, detail="Izoh yoki fayl kerak")

    comment = TaskComment(
        task_id=task_id,
        company_id=company_id,
        author_id=current_user.id,
        content=text or None,
        file_url=file_url,
        file_name=file_name,
    )
    db.add(comment)
    await db.commit()
    await db.refresh(comment)

    out = _comment_to_out(comment, current_user.full_name)
    await _broadcast_board(
        company_id,
        "task_comment",
        reason="created",
        task_id=str(task_id),
        comment=out.model_dump(mode="json"),
    )
    return out


async def _delete_comment_impl(
    db: AsyncSession, company_id: str, task_id: str, comment_id: str, current_user: User
):
    _, is_pm = await _get_visible_task(db, company_id, task_id, current_user)
    result = await db.execute(
        select(TaskComment).where(
            TaskComment.id == comment_id,
            TaskComment.task_id == task_id,
            TaskComment.company_id == company_id,
        )
    )
    comment = result.scalar_one_or_none()
    if comment is None:
        raise HTTPException(status_code=404, detail="Izoh topilmadi")
    if not is_pm and str(comment.author_id) != str(current_user.id):
        raise HTTPException(status_code=403, detail="Faqat o'z izohingizni o'chira olasiz")

    await db.delete(comment)
    await db.commit()
    await _broadcast_board(
        company_id,
        "task_comment",
        reason="deleted",
        task_id=str(task_id),
        comment_id=str(comment_id),
    )


# Preferred static "/comments/..." paths (like /history).
@router.get("/comments/{task_id}", response_model=list[TaskCommentOut])
async def list_task_comments(
    company_id: str,
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return await _list_comments_impl(db, company_id, task_id, current_user)


@router.post("/comments/{task_id}", response_model=TaskCommentOut, status_code=201)
async def create_task_comment(
    company_id: str,
    task_id: str,
    content: str = Form(""),
    file: UploadFile | None = File(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return await _create_comment_impl(db, company_id, task_id, current_user, content, file)


@router.delete("/comments/{task_id}/{comment_id}", status_code=204)
async def delete_task_comment(
    company_id: str,
    task_id: str,
    comment_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _delete_comment_impl(db, company_id, task_id, comment_id, current_user)


# Legacy aliases kept for clients still calling /{task_id}/comments.
@router.get("/{task_id}/comments", response_model=list[TaskCommentOut], include_in_schema=False)
async def list_task_comments_legacy(
    company_id: str,
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return await _list_comments_impl(db, company_id, task_id, current_user)


@router.post("/{task_id}/comments", response_model=TaskCommentOut, status_code=201, include_in_schema=False)
async def create_task_comment_legacy(
    company_id: str,
    task_id: str,
    content: str = Form(""),
    file: UploadFile | None = File(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return await _create_comment_impl(db, company_id, task_id, current_user, content, file)


@router.delete("/{task_id}/comments/{comment_id}", status_code=204, include_in_schema=False)
async def delete_task_comment_legacy(
    company_id: str,
    task_id: str,
    comment_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _delete_comment_impl(db, company_id, task_id, comment_id, current_user)


# WebSocket lives outside the company-prefixed router so the path stays stable.
ws_router = APIRouter(tags=["tasks"])


@ws_router.websocket("/ws/tasks/{company_id}")
async def tasks_board_ws(websocket: WebSocket, company_id: str):
    token = websocket.query_params.get("token")
    user_id = decode_access_token(token) if token else None
    if user_id is None:
        await websocket.close(code=4401)
        return

    async with async_session() as db:
        try:
            await _require_member(db, company_id, user_id)
        except HTTPException:
            await websocket.close(code=4403)
            return

    await task_board_manager.connect(str(company_id), websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        task_board_manager.disconnect(str(company_id), websocket)
